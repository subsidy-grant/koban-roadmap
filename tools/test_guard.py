# -*- coding: utf-8 -*-
"""出力前の自己点検（最後の関所）が働くかを確かめる。
様式の読み取りを間違えても、噛み合わない中身のまま出力しないこと。
あわせて、追加した3項目（事業形態・産業分類の番号・登記所在地の郵便番号）を確かめる。"""
import json, os, sys, warnings
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")
from playwright.sync_api import sync_playwright
import openpyxl

PAGE = "http://127.0.0.1:8935/documents.html"
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "soui")
FORMS = os.path.join(os.path.dirname(HERE), "forms")
OUT = os.path.join(HERE, "guard_out")
os.makedirs(OUT, exist_ok=True)
ok = ng = 0


def chk(cond, name, got=""):
    global ok, ng
    if cond:
        ok += 1
        print("OK ", name, ("" if got == "" else "→ " + str(got)[:44]))
    else:
        ng += 1
        print("NG ", name, "→", str(got)[:70])


def find(part):
    for root, _, names in os.walk(FORMS):
        for n in names:
            if part in n:
                return os.path.join(root, n)
    return None


def run(company, targets):
    got, logs = {}, {}
    with sync_playwright() as pw:
        b = pw.chromium.launch()
        ctx = b.new_context(viewport={"width": 1280, "height": 900}, accept_downloads=True)
        page = ctx.new_page()
        errs = []
        page.on("pageerror", lambda e: errs.append(str(e)))
        page.add_init_script("localStorage.setItem('koban_company', %s);"
                             % json.dumps(json.dumps(company, ensure_ascii=False)))
        page.goto(PAGE, wait_until="load")
        page.wait_for_timeout(1000)
        page.evaluate("document.querySelectorAll('main > section').forEach(s=>s.classList.add('is-open'))")
        page.evaluate("var d=document.getElementById('afManualBox'); if(d) d.open=true;")
        for p in targets:
            fn = os.path.basename(p)
            page.evaluate("window.KOBAN_FILL_LOG = []")
            page.set_input_files("#fileInput", p)
            page.wait_for_function("(window.KOBAN_FILL_LOG||[]).length > 0", timeout=30000)
            page.wait_for_timeout(300)
            logs[fn] = page.evaluate("window.KOBAN_FILL_LOG[0]")
            try:
                with page.expect_download(timeout=12000) as dl:
                    page.eval_on_selector("#frDone .fr-list > li:last-child a.btn-dl", "e=>e.click()")
                got[fn] = os.path.join(OUT, fn)
                dl.value.save_as(got[fn])
            except Exception:
                got[fn] = None
        b.close()
    chk(not errs, "JSエラーなし", errs[:1])
    return got, logs


APP = os.path.join(SRC, "申請書.xlsx")
DAI = find("cadf1a40")

print("=== 誤った中身を入れても、噛み合わない欄には書かない ===")
bad = {
    "entity": "法人", "name": "株式会社サンプル美容", "kana": "株式会社コバン",   # ふりがなに漢字
    "houjin": "1234567890123", "zip": "1500001", "addr": "東京都渋谷区神宮前1-2-3",
    "title": "代表取締役", "rep": "山田 太郎", "repKana": "山田 太郎",           # ふりがなに漢字
    "tel": "03-1234-5678", "mail": "example.jp",                  # @が無い
    "employees": "八人",                                                   # 数でない
    "industry": "美容業", "business": "美容室の経営",
}
got, logs = run(bad, [APP])
wb = openpyxl.load_workbook(got["申請書.xlsx"])
s1 = wb["申請者1"]
chk(not str(s1["B15"].value or "").strip(), "漢字入りのふりがなは法人フリガナ欄に入らない", s1["B15"].value)
chk(not str(s1["C20"].value or "").strip(), "漢字入りのふりがなは代表者フリガナ欄に入らない", s1["C20"].value)
chk(not str(s1["G21"].value or "").strip(), "@の無い値はメール欄に入らない", s1["G21"].value)
s2 = wb["申請者2"]
chk(not str(s2["I3"].value or "").strip() or "八" not in str(s2["I3"].value),
    "数でない人数は従業員数の欄に入らない", s2["I3"].value)
reasons = " ".join(x["reason"] for x in logs["申請書.xlsx"]["skips"])
chk("ふりがな" in reasons, "ふりがなの理由が出ている")
chk("メールアドレス" in reasons, "メールアドレスの理由が出ている")
chk("人数" in reasons or "数以外" in reasons, "人数の理由が出ている")

print("\n=== 追加した3項目 ===")
good = {
    "entity": "法人", "name": "株式会社サンプル美容", "kana": "カブシキガイシャサンプルビヨウ",
    "houjin": "1234567890123", "zip": "150-0001", "addr": "東京都渋谷区神宮前1-2-3",
    "regAddr": "東京都千代田区丸の内1-1-1", "regZip": "100-0005",
    "title": "代表取締役", "rep": "山田 太郎", "repKana": "ヤマダ タロウ",
    "tel": "03-1234-5678", "mail": "info@example.jp",
    "employees": "8", "capital": "3000000", "industry": "美容業",
    "industryCode": "78", "business": "美容室の経営", "hokenNo": "1301-123456-7",
}
tg = [APP] + ([DAI] if DAI else [])
got2, logs2 = run(good, tg)
wb2 = openpyxl.load_workbook(got2["申請書.xlsx"])
chk(wb2["表紙"]["J7"].value == "法人", "01で選んだ事業形態が使われる", wb2["表紙"]["J7"].value)
s1b = wb2["申請者1"]
chk(s1b["C29"].value == good["regAddr"], "登記上の所在地が入る", s1b["C29"].value)
chk(str(s1b["C28"].value or "").replace("　", "").find("100-0005") >= 0,
    "登記上の所在地の郵便番号が入る", s1b["C28"].value)
chk(str(s1b["C17"].value or "").replace("　", "") == "150-0001",
    "本店の郵便番号は本店の欄のまま", s1b["C17"].value)

if DAI:
    fn = os.path.basename(DAI)
    w3 = openpyxl.load_workbook(got2[fn])
    ws = w3["【代】第１号"]
    hit = [c.coordinate for row in ws.iter_rows(min_row=24, max_row=28) for c in row
           if str(c.value or "").strip() == "78"]
    chk(bool(hit), "分類番号の欄に産業分類の番号が入る", hit[:3])
    filled = " ".join(f["where"] + "=" + str(f["value"]) for f in logs2[fn]["fills"])
    chk("78" in filled, "記入の記録にも残る", [x for x in filled.split() if "78" in x][:2])

print("\n=== 個人事業主を選んだとき（名前に「株式会社」があっても選択が勝つ）===")
c3 = dict(good)
c3["entity"] = "個人事業主"
c3["houjin"] = ""
got3, _ = run(c3, [APP])
wb4 = openpyxl.load_workbook(got3["申請書.xlsx"])
chk(wb4["表紙"]["J7"].value == "個人事業主", "選んだとおり個人事業主になる", wb4["表紙"]["J7"].value)
chk(not str(wb4["申請者1"]["B15"].value or "").strip(), "法人の欄は空のまま", wb4["申請者1"]["B15"].value)

print("\n%s (OK %d / NG %d)" % ("ALL OK" if ng == 0 else "NG あり", ok, ng))
sys.exit(1 if ng else 0)
