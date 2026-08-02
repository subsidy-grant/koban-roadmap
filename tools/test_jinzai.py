# -*- coding: utf-8 -*-
"""人材開発支援助成金の様式（第1-1号・第4-1号）で、2026-08-03に見つかった
5つの取りこぼしが再発しないことを確かめる。

 ・雇用保険適用事業所番号が入る（見出しの効き目が下まで届きすぎて空欄だった）
 ・その番号は「4桁-6桁-1桁」に分けて入る
 ・事業所の所在地は、郵便番号を枠に分け、住所は次の行に入る
 ・担当者の所属・役職に代表者の役職が入る（担当者＝代表者のとき）
 ・「④MAIL」と書かれた欄にもメールが入る（「④メール」だけ入っていた）
 ・法人番号の欄は、結合が抜けていても読める幅にしてから入れる
"""
import json, os, re, sys, warnings
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")
import openpyxl
from playwright.sync_api import sync_playwright

HERE = os.path.dirname(os.path.abspath(__file__))
FORMS = os.path.join(os.path.dirname(HERE), "forms")
OUT = os.path.join(HERE, "jinzai_out")
os.makedirs(OUT, exist_ok=True)
PAGE = "http://127.0.0.1:8935/documents.html"
C = {"entity": "法人", "name": "株式会社サンプル美容", "kana": "カブシキガイシャサンプルビヨウ",
     "houjin": "1234567890123", "zip": "150-0001", "addr": "東京都渋谷区神宮前1-2-3",
     "title": "代表取締役", "rep": "山田 太郎", "repKana": "ヤマダ タロウ",
     "tel": "03-1234-5678", "mail": "info@example.jp",
     "employees": "8", "capital": "3000000", "industry": "美容業", "industryCode": "78",
     "business": "美容室の経営", "staff": "山田 太郎", "staffTel": "03-1234-5678",
     "hokenNo": "1301-123456-7"}
ok = ng = 0


def chk(cond, name, got=""):
    global ok, ng
    if cond:
        ok += 1
        print("OK ", name, ("" if got == "" else "→ " + str(got)[:42]))
    else:
        ng += 1
        print("NG ", name, "→", str(got)[:70])


paths = {}
for root, _, names in os.walk(FORMS):
    for n in names:
        paths[n] = os.path.join(root, n)
A = "6adefe11_001698095.xlsx"   # 様式第1-1号 職業訓練実施計画届
B = "b93f369a_001690646.xlsx"   # 様式第4-1号 支給申請書
if A not in paths or B not in paths:
    print("様式が見つかりません（forms/ を確認）"); sys.exit(1)

with sync_playwright() as pw:
    b = pw.chromium.launch()
    ctx = b.new_context(viewport={"width": 1280, "height": 900}, accept_downloads=True)
    page = ctx.new_page()
    errs = []
    page.on("pageerror", lambda e: errs.append(str(e)))
    page.add_init_script("localStorage.setItem('koban_company', %s);"
                         % json.dumps(json.dumps(C, ensure_ascii=False)))
    page.goto(PAGE, wait_until="load")
    page.wait_for_timeout(1000)
    page.evaluate("document.querySelectorAll('main > section').forEach(s=>s.classList.add('is-open'))")
    page.evaluate("var d=document.getElementById('afManualBox'); if(d) d.open=true;")
    for fn in (A, B):
        page.evaluate("window.KOBAN_FILL_LOG = []")
        page.set_input_files("#fileInput", paths[fn])
        page.wait_for_function("(window.KOBAN_FILL_LOG||[]).length > 0", timeout=30000)
        page.wait_for_timeout(300)
        with page.expect_download(timeout=15000) as dl:
            page.eval_on_selector("#frDone .fr-list > li:last-child a.btn-dl", "e=>e.click()")
        dl.value.save_as(os.path.join(OUT, fn))
    b.close()
chk(not errs, "JSエラーなし", errs[:1])


def val(ws, ref):
    return str(ws[ref].value or "").strip()


print("\n=== 様式第1-1号（職業訓練実施計画届）===")
w1 = openpyxl.load_workbook(os.path.join(OUT, A)).worksheets[0]
chk(val(w1, "AF14") == C["houjin"], "法人番号が入る", val(w1, "AF14"))
chk(val(w1, "K26") == C["name"], "雇用保険適用事業所の名称", val(w1, "K26"))
chk([val(w1, "AN26"), val(w1, "AS26"), val(w1, "AZ26")] == ["1301", "123456", "7"],
    "事業所番号が4桁-6桁-1桁に分かれて入る",
    [val(w1, "AN26"), val(w1, "AS26"), val(w1, "AZ26")])
chk(val(w1, "M27") == "150" and val(w1, "Q27") == "0001",
    "事業所所在地の郵便番号が枠に分かれて入る", [val(w1, "M27"), val(w1, "Q27")])
chk(val(w1, "K28") == C["addr"], "住所は次の行の欄に入る", val(w1, "K28"))
chk("〒" not in val(w1, "K27") or val(w1, "K27") == "（〒",
    "郵便番号の枠じたいは壊さない", val(w1, "K27"))
chk(val(w1, "R29") == C["rep"], "担当者氏名", val(w1, "R29"))
chk(val(w1, "AM29") == C["title"], "担当者の所属・役職に代表者の役職", val(w1, "AM29"))
chk(val(w1, "AM30") == C["mail"], "「④MAIL」の欄にメールが入る", val(w1, "AM30"))
chk(not val(w1, "AF17") and not val(w1, "AF19") and not val(w1, "AF20"),
    "代理人・社労士の欄は空のまま", [val(w1, "AF17"), val(w1, "AF19"), val(w1, "AF20")])

print("\n=== 様式第4-1号（支給申請書）===")
wb2 = openpyxl.load_workbook(os.path.join(OUT, B))
w2 = wb2.worksheets[0]
chk(val(w2, "AF14") == C["houjin"], "法人番号が入る", val(w2, "AF14"))
mg = [str(m) for m in w2.merged_cells.ranges if m.min_row == 14 and m.min_col >= 32]
chk(any(m.startswith("AF14:") and m != "AF14:AF14" for m in mg),
    "法人番号の欄が読める幅に結合される（配布元は結合漏れ）", mg[:3])
chk([val(w2, "AN30"), val(w2, "AS30"), val(w2, "AZ30")] == ["1301", "123456", "7"],
    "事業所番号が4桁-6桁-1桁に分かれて入る",
    [val(w2, "AN30"), val(w2, "AS30"), val(w2, "AZ30")])
chk(val(w2, "AM31") == C["title"], "担当者の所属・役職に代表者の役職", val(w2, "AM31"))
chk(val(w2, "AM32") == C["mail"], "「④メール」の欄にメールが入る", val(w2, "AM32"))
chk(val(w2, "K27") == C["industry"], "主たる事業", val(w2, "K27"))

print("\n=== 担当者が代表者と別の方のとき ===")
C2 = dict(C)
C2["staff"] = "鈴木 花子"
with sync_playwright() as pw:
    b = pw.chromium.launch()
    ctx = b.new_context(viewport={"width": 1280, "height": 900}, accept_downloads=True)
    page = ctx.new_page()
    page.add_init_script("localStorage.setItem('koban_company', %s);"
                         % json.dumps(json.dumps(C2, ensure_ascii=False)))
    page.goto(PAGE, wait_until="load")
    page.wait_for_timeout(1000)
    page.evaluate("document.querySelectorAll('main > section').forEach(s=>s.classList.add('is-open'))")
    page.evaluate("var d=document.getElementById('afManualBox'); if(d) d.open=true;")
    page.set_input_files("#fileInput", paths[A])
    page.wait_for_function("(window.KOBAN_FILL_LOG||[]).length > 0", timeout=30000)
    page.wait_for_timeout(300)
    with page.expect_download(timeout=15000) as dl:
        page.eval_on_selector("#frDone .fr-list > li:last-child a.btn-dl", "e=>e.click()")
    dl.value.save_as(os.path.join(OUT, "staff_" + A))
    b.close()
w3 = openpyxl.load_workbook(os.path.join(OUT, "staff_" + A)).worksheets[0]
chk(val(w3, "R29") == "鈴木 花子", "担当者氏名はその方の名前", val(w3, "R29"))
chk(not val(w3, "AM29"), "別の方なら所属・役職は空欄（役職を預かっていない）", val(w3, "AM29"))

print("\n%s (OK %d / NG %d)" % ("ALL OK" if ng == 0 else "NG あり", ok, ng))
sys.exit(1 if ng else 0)
