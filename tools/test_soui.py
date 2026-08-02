# -*- coding: utf-8 -*-
"""2026-08-03に実物で見つけた記入の誤りが、二度と起きないことを確かめる。
対象：東京都中小企業振興公社「創意工夫チャレンジ促進事業」の実物3様式。
 ・フリガナの欄に漢字を入れない
 ・代理申請者の欄には何も入れない
 ・役職と氏名が別の欄なら分けて入れる（案内の「（役職）」を潰さない）
 ・法人と個人事業主で分かれた欄は、当てはまるほうだけ
 ・都内登記所在地の欄に本店の住所を入れない
 ・文字を幅に押し込む指定を値に持ち込まない
"""
import json, os, sys, warnings, zipfile
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")
from playwright.sync_api import sync_playwright
from xml.dom import minidom
import openpyxl

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
PAGE = "http://127.0.0.1:8935/documents.html"
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "soui")
OUT = os.path.join(HERE, "soui_test")
os.makedirs(OUT, exist_ok=True)
BASE = {
    "name": "株式会社サンプル美容", "kana": "カブシキガイシャサンプルビヨウ", "houjin": "7000012050002",
    "zip": "150-0001", "addr": "東京都渋谷区神宮前1-2-3",
    "title": "代表取締役", "rep": "山田 太郎", "repKana": "タイラ ヒロシ",
    "tel": "03-1234-5678", "mail": "info@example.jp",
    "employees": "8", "capital": "3000000", "founded": "2019-04-01",
    "industry": "美容業", "business": "美容室の経営",
    "staff": "山田 太郎", "staffTel": "03-1234-5678", "hokenNo": "1301-123456-7",
}
ok = ng = 0


def chk(cond, name, got=""):
    global ok, ng
    if cond:
        ok += 1
        print("OK ", name, ("" if got == "" else "→ " + str(got)[:40]))
    else:
        ng += 1
        print("NG ", name, "→", str(got)[:60])


def text_of(el):
    out = []
    for nd in el.getElementsByTagNameNS(W, "*"):
        if nd.localName == "t":
            out.append(nd.firstChild.nodeValue if nd.firstChild else "")
        elif nd.localName == "noBreakHyphen":
            out.append("-")
    return "".join(out)


def run(company, targets):
    """様式に記入させて、記入後のファイルの場所を返す"""
    got = {}
    with sync_playwright() as pw:
        b = pw.chromium.launch()
        ctx = b.new_context(viewport={"width": 1280, "height": 900}, accept_downloads=True)
        page = ctx.new_page()
        errs = []
        page.on("pageerror", lambda e: errs.append(str(e)))
        page.add_init_script("localStorage.setItem('koban_company', %s);"
                             % json.dumps(json.dumps(company, ensure_ascii=False)))
        page.goto(PAGE, wait_until="load")
        page.wait_for_timeout(1000)
        page.evaluate("document.querySelectorAll('main > section').forEach(s=>s.classList.add('is-open'))")
        page.evaluate("var d=document.getElementById('afManualBox'); if(d) d.open=true;")
        for fn in targets:
            page.evaluate("window.KOBAN_FILL_LOG = []")
            page.set_input_files("#fileInput", os.path.join(SRC, fn))
            page.wait_for_function("(window.KOBAN_FILL_LOG||[]).length > 0", timeout=30000)
            page.wait_for_timeout(300)
            with page.expect_download(timeout=15000) as dl:
                page.eval_on_selector("#frDone .fr-list > li:last-child a.btn-dl", "e=>e.click()")
            dest = os.path.join(OUT, fn)
            dl.value.save_as(dest)
            got[fn] = dest
        b.close()
    chk(not errs, "JSエラーなし", errs[:1])
    return got


print("=== 会社情報に「登記上の所在地」の入力欄がある ===")
with sync_playwright() as pw:
    b = pw.chromium.launch()
    pg = b.new_context().new_page()
    pg.goto(PAGE, wait_until="load")
    pg.wait_for_timeout(900)
    chk(pg.locator("#f_regAddr").count() == 1, "登記上の所在地の入力欄がある")
    lab = pg.evaluate("(document.querySelector('#f_kana')||{}).previousElementSibling ?"
                      " document.querySelector('#f_kana').parentElement.textContent : ''")
    chk("フリガナ" in lab, "法人名のフリガナの入力欄がある", lab.strip()[:24])
    b.close()

APP = "申請書.xlsx"
SEI = "誓約書（助成金申請に関する誓約書）.xlsx"
DOU = "同意書（代理申請用）.docx"

print("\n=== 登記所在地は未入力（本店＝都内）===")
out = run(BASE, [APP, SEI, DOU])

wb = openpyxl.load_workbook(out[APP])
hy = wb["表紙"]
chk(hy["J7"].value == "法人", "表紙：事業形態が「法人」になる", hy["J7"].value)
chk(str(hy["J13"].value or "") == "（役職）", "表紙：案内の「（役職）」を消していない", hy["J13"].value)
chk(str(hy["J14"].value or "") == "（氏名）", "表紙：案内の「（氏名）」を消していない", hy["J14"].value)
chk(hy["L13"].value == "代表取締役", "表紙：役職は役職の欄に入る", hy["L13"].value)
chk(hy["L14"].value == "山田 太郎", "表紙：氏名は氏名の欄に入る", hy["L14"].value)

s1 = wb["申請者1"]
chk(s1["B15"].value == BASE["kana"], "申請者1：法人のフリガナに会社のふりがな", s1["B15"].value)
chk(s1["C20"].value == BASE["repKana"], "申請者1：代表者のフリガナに人のふりがな", s1["C20"].value)
chk(not str(s1["B8"].value or "").strip(), "申請者1：個人事業主の欄は空のまま", s1["B8"].value)
chk(not str(s1["B11"].value or "").strip() or "〒" in str(s1["B11"].value),
    "申請者1：個人事業主の郵便番号は空のまま", s1["B11"].value)
chk(not str(s1["C29"].value or "").strip(), "申請者1：都内登記所在地に本店の住所を入れない", s1["C29"].value)
chk(str(s1["C17"].value or "").replace("　", "") == "150-0001", "申請者1：法人の郵便番号は入る", s1["C17"].value)
kanji = [c for c in "山田 太郎" ]
for ref in ("B8", "C20", "B15"):
    v = str(s1[ref].value or "")
    chk(not any(k in v for k in kanji) or ref == "X", "申請者1：%s のフリガナ欄に漢字が無い" % ref, v)

wb2 = openpyxl.load_workbook(out[SEI])
ws2 = wb2.worksheets[0]
chk(ws2["B37"].value == BASE["name"], "誓約書：事業者名称に会社名が入る", ws2["B37"].value)
chk(ws2["B38"].value == BASE["rep"], "誓約書：代表者氏名が入る", ws2["B38"].value)
chk(not str(ws2["B36"].value or "").strip(), "誓約書：提出日は空のまま（本人が書く欄）", ws2["B36"].value)

doc = minidom.parseString(zipfile.ZipFile(out[DOU]).read("word/document.xml"))
tbls = [t for t in doc.getElementsByTagNameNS(W, "tbl")]
chk(len(tbls) >= 2, "同意書：表が2つある", len(tbls))
agent_vals, self_rows = [], {}
for ti, tbl in enumerate(tbls[:2]):
    for tr in tbl.getElementsByTagNameNS(W, "tr"):
        if tr.parentNode is not tbl:
            continue
        tcs = [tc for tc in tr.getElementsByTagNameNS(W, "tc") if tc.parentNode is tr]
        lab = text_of(tcs[0]).replace("　", " ").strip()
        val = text_of(tcs[1]).strip() if len(tcs) > 1 else ""
        if ti == 0:
            agent_vals.append((lab, val))
        else:
            self_rows[lab] = val
chk(all(not v for _, v in agent_vals), "同意書：代理申請者の欄は全部空のまま",
    [x for x in agent_vals if x[1]])
chk(self_rows.get("法人名又は屋号") == BASE["name"], "同意書：申請者の法人名が入る", self_rows.get("法人名又は屋号"))
chk(self_rows.get("電話番号：本社") == BASE["tel"], "同意書：本社の電話が入る", self_rows.get("電話番号：本社"))
chk(self_rows.get("電話番号：連絡担当者") == BASE["staffTel"], "同意書：連絡担当者の電話が入る",
    self_rows.get("電話番号：連絡担当者"))
chk(not self_rows.get("都内登記所在地"), "同意書：登記所在地は未入力なら空のまま",
    self_rows.get("都内登記所在地"))
# 文字を幅に押し込む指定を持ち込んでいないか
squeeze = []
for tbl in tbls[1:2]:
    for tr in tbl.getElementsByTagNameNS(W, "tr"):
        tcs = [tc for tc in tr.getElementsByTagNameNS(W, "tc") if tc.parentNode is tr]
        if len(tcs) < 2 or not text_of(tcs[1]).strip():
            continue
        for rPr in tcs[1].getElementsByTagNameNS(W, "rPr"):
            for ch in rPr.childNodes:
                if ch.nodeType == 1 and ch.localName in ("fitText", "spacing", "w", "kern"):
                    squeeze.append(text_of(tcs[0]).strip()[:12] + ":" + ch.localName)
chk(not squeeze, "同意書：値に文字の押し込み指定を持ち込まない", squeeze[:3])
# 見出しの段落に書き足していないか
heads = []
body = doc.getElementsByTagNameNS(W, "body")[0]
for node in body.childNodes:
    if node.nodeType == 1 and node.localName == "p":
        t = text_of(node).strip()
        if t.startswith("〇") and BASE["name"] in t:
            heads.append(t)
chk(not heads, "同意書：表の見出しの行に会社名を書き足さない", heads[:2])

print("\n=== 登記所在地を入れた場合 ===")
c2 = dict(BASE)
c2["regAddr"] = "東京都千代田区丸の内1-1-1"
out2 = run(c2, [APP, DOU])
wb3 = openpyxl.load_workbook(out2[APP])
s1b = wb3["申請者1"]
chk(s1b["C29"].value == c2["regAddr"], "申請者1：都内登記所在地に登記の住所が入る", s1b["C29"].value)
chk(s1b["C18"].value in (None, "") or "丸の内" not in str(s1b["C18"].value),
    "申請者1：法人の住所欄は数式のまま", s1b["C18"].value)
doc2 = minidom.parseString(zipfile.ZipFile(out2[DOU]).read("word/document.xml"))
rows2 = {}
for tbl in list(doc2.getElementsByTagNameNS(W, "tbl"))[1:2]:
    for tr in tbl.getElementsByTagNameNS(W, "tr"):
        tcs = [tc for tc in tr.getElementsByTagNameNS(W, "tc") if tc.parentNode is tr]
        if len(tcs) > 1:
            rows2[text_of(tcs[0]).replace("　", " ").strip()] = text_of(tcs[1]).strip()
chk(rows2.get("都内登記所在地") == c2["regAddr"], "同意書：都内登記所在地に登記の住所が入る",
    rows2.get("都内登記所在地"))
chk(rows2.get("本社所在地") == BASE["addr"], "同意書：本社所在地はそのまま", rows2.get("本社所在地"))

print("\n=== 個人事業主のとき（法人番号なし・屋号だけ）===")
c3 = {"name": "ヘアサロンこばん", "kana": "ヘアサロンコバン", "zip": "150-0001",
      "addr": "東京都渋谷区神宮前1-2-3", "rep": "山田 太郎", "repKana": "タイラ ヒロシ",
      "tel": "03-1234-5678", "mail": "info@example.jp",
      "employees": "3", "industry": "美容業", "business": "美容室の経営"}
out3 = run(c3, [APP])
wb4 = openpyxl.load_workbook(out3[APP])
chk(wb4["表紙"]["J7"].value == "個人事業主", "表紙：事業形態が「個人事業主」になる", wb4["表紙"]["J7"].value)
s1c = wb4["申請者1"]
# A8「フリガナ」の1行下は A9「申請者氏名」。つまり本人の氏名のふりがなの欄
chk(str(s1c["B8"].value or "") == c3["repKana"], "申請者1：個人事業主のフリガナは本人のふりがな", s1c["B8"].value)
chk(not str(s1c["B15"].value or "").strip(), "申請者1：法人の欄は空のまま", s1c["B15"].value)
chk(not str(s1c["C20"].value or "").strip(), "申請者1：法人代表者のフリガナも空のまま", s1c["C20"].value)

print("\n%s (OK %d / NG %d)" % ("ALL OK" if ng == 0 else "NG あり", ok, ng))
sys.exit(1 if ng else 0)
