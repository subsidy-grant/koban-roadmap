# -*- coding: utf-8 -*-
"""全様式・全シートを見て、「空欄の隣にある欄の名前」のうち
当サイトが読めていないものを洗い出す。辞書の穴を見つけるため。"""
import io, json, os, re, sys, warnings, zipfile
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")
import openpyxl
from xml.dom import minidom
from playwright.sync_api import sync_playwright

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
HERE = os.path.dirname(os.path.abspath(__file__))
FORMS = os.path.join(os.path.dirname(HERE), "forms")
SOUI = os.path.join(HERE, "soui")
PAGE = "http://127.0.0.1:8935/documents.html"

files = []
for root, _, names in os.walk(FORMS):
    for n in names:
        if re.search(r"\.(docx|xlsx)$", n, re.I) and not n.startswith("~$"):
            files.append(os.path.join(root, n))
for n in os.listdir(SOUI):
    if re.search(r"\.(docx|xlsx)$", n, re.I):
        files.append(os.path.join(SOUI, n))
print("見る様式:", len(files), "件")


def text_of(el):
    out = []
    for nd in el.getElementsByTagNameNS(W, "*"):
        if nd.localName == "t":
            out.append(nd.firstChild.nodeValue if nd.firstChild else "")
    return "".join(out)


def looks_label(t):
    t = (t or "").strip()
    if not t or len(t) > 22:
        return False
    if re.fullmatch(r"[\d\s　,.\-−ー－()（）%％円人年月日〒:：/／]+", t):
        return False
    return True


cands = {}   # 欄の名前 -> [(ファイル, 場所)]


def add(name, where):
    key = re.sub(r"[\s　]+", "", name)
    if not key:
        return
    cands.setdefault(key, []).append(where)


for p in files:
    base = os.path.basename(p)
    try:
        if p.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(p)
            for ws in wb.worksheets:
                if ws.max_row > 400:
                    continue
                for row in ws.iter_rows(max_row=min(ws.max_row, 200), max_col=min(ws.max_column, 40)):
                    for i, c in enumerate(row):
                        v = c.value
                        if not isinstance(v, str) or not looks_label(v):
                            continue
                        nxt = row[i + 1] if i + 1 < len(row) else None
                        nv = nxt.value if nxt is not None else None
                        if nv is None or (isinstance(nv, str) and not nv.strip()):
                            add(v, "%s／%s／%s" % (base, ws.title, c.coordinate))
        else:
            doc = minidom.parseString(zipfile.ZipFile(p).read("word/document.xml"))
            for tbl in doc.getElementsByTagNameNS(W, "tbl"):
                for tr in tbl.getElementsByTagNameNS(W, "tr"):
                    if tr.parentNode is not tbl:
                        continue
                    tcs = [tc for tc in tr.getElementsByTagNameNS(W, "tc") if tc.parentNode is tr]
                    for i, tc in enumerate(tcs[:-1]):
                        t = text_of(tc).strip()
                        if not looks_label(t):
                            continue
                        if not text_of(tcs[i + 1]).strip():
                            add(t, "%s／表" % base)
    except Exception as e:
        print("  読めず:", base, str(e)[:60])

print("空欄の隣にある欄の名前:", len(cands), "種類")

names = sorted(cands.keys())
with sync_playwright() as pw:
    b = pw.chromium.launch()
    page = b.new_context().new_page()
    page.goto(PAGE, wait_until="load")
    page.wait_for_timeout(1200)
    got = page.evaluate("ns => ns.map(n => [n, window.KOBAN_MATCH_LABEL(n)])", names)
    b.close()

unknown = [(n, k) for n, k in got if not k]
known = [(n, k) for n, k in got if k]
print("読めている:", len(known), "／ 読めていない:", len(unknown))

# よく出るものから並べる
unknown.sort(key=lambda x: -len(cands[x[0]]))
print("\n=== 読めていない欄の名前（出てくる回数が多い順・上位60）===")
for n, _ in unknown[:60]:
    print("  %-3d %-30s %s" % (len(cands[n]), n[:30], cands[n][0][:52]))

io.open(os.path.join(HERE, "sweep_labels.json"), "w", encoding="utf-8").write(
    json.dumps({"unknown": [[n, len(cands[n]), cands[n][:3]] for n, _ in unknown],
                "known": [[n, k] for n, k in known]}, ensure_ascii=False, indent=1))
print("\n明細: sweep_labels.json")
