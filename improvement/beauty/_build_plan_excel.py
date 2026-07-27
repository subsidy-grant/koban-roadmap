# 事業計画書 Excel版ジェネレーター
# 入力: _plan_data.json (_build_plans.js が書き出す。plan-XX.htmlと同じ数値データ)
# 出力: plan-01.xlsx 〜 plan-10.xlsx (各10シート = HTML版の10ページに対応、数値・文章を直接編集可能)
# 実行: python3 _build_plan_excel.py
#
# 旧 beauty-ai-factory プロジェクト(2026-07-18作成)から移行。
# _build_plans.js → _build_plans_v2.js → 本スクリプト の順に実行し、HTML版と数値を一致させること。

import json
import math
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

CHART_COLORS = ["C9B79C", "3E6B4F", "B98A2F", "8A5A2B", "A4453C"]
ROW_CM = 0.53  # 既定の行高(15pt)をcm換算した目安値。チャートの縦幅から必要行数を逆算するために使う
PRINT_LAST_COL = "E"  # 全シート共通のコンテンツ幅(A〜E)。印刷範囲の右端に使う
# 結合セルの幅(span=結合する列数, A=22/B-E=15の列幅からpx換算→cm換算した目安値)
SPAN_WIDTH_CM = {2: 7.12, 3: 10.03, 4: 12.94, 5: 15.86}
COL_WIDTH_CM = {1: 4.207, 2: 2.91, 3: 2.91, 4: 2.91, 5: 2.91}  # 単一列(A/B-E)の幅


def estimate_height_from_width(text, width_cm, font_pt=10):
    """折り返し行数からその行に必要な高さ(pt)を概算し、セルの文字が見切れないようにする"""
    char_w_cm = font_pt / 72 * 2.54 * 0.98  # 全角文字1つ分の目安幅
    chars_per_line = max(6, int(width_cm / char_w_cm))
    lines = max(1, math.ceil(len(str(text)) / chars_per_line))
    line_h_pt = font_pt * 1.3
    return max(15, lines * line_h_pt + 4)


def estimate_row_height(text, span, font_pt=10):
    """結合セル(put_lead/put_callout/put_noteが使う)向けの行高estimate"""
    return estimate_height_from_width(text, SPAN_WIDTH_CM.get(span, 12.94), font_pt)


def row_height_for_cells(cells, font_pt=10):
    """1行の各セル(列番号, テキスト)から、その行に必要な最大高さを求める(表の行の見切れ防止)"""
    h = 15
    for col_idx, text in cells:
        w = COL_WIDTH_CM.get(col_idx, 2.91)
        h = max(h, estimate_height_from_width(text, w, font_pt))
    return h

HERE = os.path.dirname(os.path.abspath(__file__))

INK = "26221E"
SUB = "6F675E"
ACCENT = "8A5A2B"
GREEN = "3E6B4F"
GOLD = "B98A2F"
LINE_BG = "EFE9E0"

TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color=ACCENT)
H2_FONT = Font(name="Yu Gothic", size=11, bold=True, color=ACCENT)
LABEL_FONT = Font(name="Yu Gothic", size=10, bold=True, color=SUB)
BODY_FONT = Font(name="Yu Gothic", size=10, color=INK)
NOTE_FONT = Font(name="Yu Gothic", size=9, italic=True, color=SUB)
HEADER_FILL = PatternFill("solid", fgColor=LINE_BG)
THIN = Side(style="thin", color="D9D2C8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_H = Alignment(wrap_text=True, vertical="center", horizontal="center")


def sheet(wb, name):
    ws = wb.create_sheet(title=name[:31])
    ws.sheet_view.showGridLines = False
    # A4縦1ページに収まるよう、以前より控えめな幅にする(A〜Eの合計が印刷可能幅に収まる程度)
    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 15
    return ws


def apply_print_setup(ws, last_row, last_col=PRINT_LAST_COL):
    """1シート=1ページに収まるよう印刷設定(A4縦・1ページ拡大縮小・余白最小)を適用する"""
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.4, bottom=0.4, header=0.1, footer=0.1)
    ws.print_area = f"A1:{last_col}{last_row}"


def put_title(ws, row, text, page_no):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=4, value=f"{page_no} / 10").font = NOTE_FONT
    return row + 2


def put_h3(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = H2_FONT
    return row + 1


def put_lead(ws, row, text, span=4):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BODY_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = estimate_row_height(text, span, font_pt=10)
    return row + 1


def put_kv_table(ws, row, rows, header=None):
    if header:
        for j, h in enumerate(header):
            c = ws.cell(row=row, column=1 + j, value=h)
            c.font = LABEL_FONT
            c.fill = HEADER_FILL
            c.border = BORDER
        row += 1
    for k, v in rows:
        ck = ws.cell(row=row, column=1, value=k)
        ck.font = LABEL_FONT
        ck.border = BORDER
        ck.alignment = WRAP
        cv = ws.cell(row=row, column=2, value=v)
        cv.font = BODY_FONT
        cv.border = BORDER
        cv.alignment = WRAP
        ws.row_dimensions[row].height = row_height_for_cells([(1, k), (2, v)])
        row += 1
    return row + 1


def put_data_table(ws, row, headers, rows):
    header_cells = []
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = LABEL_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = WRAP_H
        header_cells.append((1 + j, h))
    ws.row_dimensions[row].height = row_height_for_cells(header_cells)
    row += 1
    for r in rows:
        cells = []
        for j, v in enumerate(r):
            c = ws.cell(row=row, column=1 + j, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = WRAP
            cells.append((1 + j, v))
        ws.row_dimensions[row].height = row_height_for_cells(cells)
        row += 1
    return row + 1


def put_note(ws, row, text, span=4):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = estimate_row_height(text, span, font_pt=9)
    return row + 2


def add_bar_chart(ws, row, title, header_row, last_row, cat_col=1, data_col_start=2, data_col_end=3,
                   width=15.5, height=6):
    """表のすぐ下(A列位置・コンテンツ幅いっぱい)にグラフを配置し、専有した行数だけ進めた次の行を返す。
    header_row: 見出し行(指標名列+系列名の行) / last_row: データ最終行"""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.width = width
    chart.height = height
    chart.legend.position = "b"
    data = Reference(ws, min_col=data_col_start, max_col=data_col_end, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for i, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = CHART_COLORS[i % len(CHART_COLORS)]
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    ws.add_chart(chart, f"A{row}")
    return row + math.ceil(height / ROW_CM) + 1


def add_pie_chart(ws, row, title, header_row, last_row, cat_col=1, data_col=2, width=15.5, height=6):
    chart = PieChart()
    chart.title = title
    chart.width = width
    chart.height = height
    chart.legend.position = "r"
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showPercent = True
    ws.add_chart(chart, f"A{row}")
    return row + math.ceil(height / ROW_CM) + 1


def put_callout(ws, row, label, text, span=4):
    full = f"【{label}】{text}"
    c = ws.cell(row=row, column=1, value=full)
    c.font = BODY_FONT
    c.alignment = WRAP
    c.fill = PatternFill("solid", fgColor="F7F3EC")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = estimate_row_height(full, span, font_pt=10)
    return row + 2


def build_plan_workbook(p, biz):
    s = p["scheme_info"]
    no = p["no"]
    wb = Workbook()
    wb.remove(wb.active)

    # --- 1. 事業概要 ---
    ws = sheet(wb, "1_事業概要")
    r = put_title(ws, 1, "1. 事業概要", 1)
    ws.cell(row=r, column=1, value=p["title"]).font = Font(name="Yu Gothic", size=13, bold=True, color=INK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 30
    r += 1
    r = put_lead(ws, r, p["sub"])
    r += 1
    r = put_h3(ws, r, "申請者概要")
    r = put_kv_table(ws, r, [
        ["事業者名", f"{biz['name']}(デモ)"], ["業種", biz["type"]], ["所在地", biz["area"]],
        ["従業員数", f"{biz['staff']}名"], ["創業", biz["founded"]], ["年商", biz["sales"]],
    ])
    r = put_h3(ws, r, "本事業のサマリー(数値は編集可能)")
    r = put_kv_table(ws, r, [
        ["投資総額(万円)", p["inv"]], ["補助率", s["rate"]], ["補助見込(万円)", s["subAmt"]],
        ["自己負担(万円)", s["self"]], ["省力化効果(時間/月)", p["saveH"]], ["投資回収(年)", p["roi"]],
    ])
    r = put_h3(ws, r, "主要KPI(導入前→導入後)")
    kpi_header = r
    r = put_data_table(ws, r, ["指標", "導入前", "導入後"],
                        [[k["label"], k["before"], k["after"]] for k in p["kpis"]])
    r = add_bar_chart(ws, r, "主要KPI(導入前→導入後)", kpi_header, kpi_header + len(p["kpis"]))
    r = put_lead(ws, r, p["problemLead"])
    r = put_callout(ws, r, "本事業の狙い",
                     f"{p['sub']}。省力化により生み出した時間を、より付加価値の高い接客・技術・提案へ再配分し、"
                     "一人当たり労働生産性と従業員の処遇改善を同時に実現する。")
    apply_print_setup(ws, r)

    # --- 2. 現状の課題と労働実態 ---
    ws = sheet(wb, "2_現状の課題")
    r = put_title(ws, 1, "2. 現状の課題と労働実態", 2)
    r = put_lead(ws, r, p["problemLead"])
    r = put_h3(ws, r, "現状の労働実態(課題の定量化)")
    r = put_data_table(ws, r, ["項目", "現状"], [
        ["対象業務の月間工数", f"約{p['saveH'] + 2}時間"],
        ["担当", "店長・スタイリストが兼務"],
        ["発生タイミング", "営業時間中/閉店後の残業"],
        ["属人性", "高(特定者に依存)"],
        ["ミス・機会損失", "恒常的に発生"],
    ])
    r = put_h3(ws, r, "省力化・属人化指標(%)")
    pct_header = r
    pct_saving = round(p["saveH"] / (p["saveH"] + 2) * 100)
    r = put_data_table(ws, r, ["指標", "%"], [["省力化率", pct_saving], ["属人業務率", 78]])
    r = add_bar_chart(ws, r, "省力化・属人化指標(%)", pct_header, pct_header + 2, data_col_end=2)
    r = put_note(ws, r, "※ 本計画の数値は業界実態に基づくデモ用の仮置きです。実申請時には自社の勤怠記録・POS実績等の裏付けデータに差し替えます。")
    apply_print_setup(ws, r)

    # --- 3. 導入システムと省力化プロセス ---
    ws = sheet(wb, "3_システム導入")
    r = put_title(ws, 1, "3. 導入システムと省力化プロセス", 3)
    r = put_lead(ws, r, f"本事業では、{p['title'].split('による')[0]}を導入する。人手で行ってきた一連の業務をAIが自動実行し、担当者は最終確認・承認のみを行う体制へ移行する。")
    r = put_h3(ws, r, "導入フロー(STEP1〜STEP5)")
    r = put_data_table(ws, r, ["STEP", "内容"], [[f"STEP{i+1}", step.replace("\n", " ")] for i, step in enumerate(p["flow"])])
    r = put_h3(ws, r, "Before(人手) / After(AI自動化)")
    r = put_data_table(ws, r, ["Before:現行フロー(人手)", "After:AI自動化フロー"], [
        ["担当者が手作業で全工程を実施", "AIが自動実行、担当は承認のみ"],
        ["作業は営業時間・閉店後を圧迫", "24時間・無人でも業務が進行"],
        ["品質・スピードが担当者に依存", "品質を標準化・均一化"],
        ["記録・転記漏れやミスが発生", "記録は自動化しミスを排除"],
    ])
    r = put_callout(ws, r, "実機デモ", f"本システムの動作は付属のプロトタイプ(prototypes/{p['proto']})で確認できる。実際の画面遷移・AI出力・効果数値を再現しており、審査における実現可能性の裏付けとする。")
    r = put_callout(ws, r, "差別化ポイント", p["distinct"])
    apply_print_setup(ws, r)

    # --- 4. 労働生産性向上の定量根拠 ---
    ws = sheet(wb, "4_生産性")
    r = put_title(ws, 1, "4. 労働生産性向上の定量根拠", 4)
    r = put_h3(ws, r, "導入前後の業務指標比較(数値)")
    kpi4_header = r
    r = put_data_table(ws, r, ["指標", "導入前", "導入後"],
                        [[k["label"], k["before"], k["after"]] for k in p["kpis"]])
    r = add_bar_chart(ws, r, "導入前後の業務指標比較", kpi4_header, kpi4_header + len(p["kpis"]))
    r = put_h3(ws, r, "詳細効果(指標別)")
    r = put_data_table(ws, r, ["指標", "導入前", "導入後", "改善"], p["effectRows"])
    r = put_callout(ws, r, "労働生産性の算定",
                     f"月間{p['saveH']}時間の省力化を年換算すると約{p['saveH']*12}時間。これを平均時間単価で換算し、"
                     f"生み出した時間の付加価値創出への再配分と合わせ、一人当たり付加価値額の年平均成長率+{p['growth_pct']}%以上を見込む。")
    r = put_callout(ws, r, "売上・利益への波及",
                     f"削減時間の高付加価値業務への再配分と、本システムによる直接的な売上効果(機会損失の回収・客単価向上等)により、年間約{p['saveY']}万円の収益改善を見込む。")
    apply_print_setup(ws, r)

    # --- 5. 賃上げ計画 ---
    ws = sheet(wb, "5_賃上げ計画")
    r = put_title(ws, 1, "5. 賃上げ計画(処遇改善)", 5)
    r = put_callout(ws, r, f"{s['badge']}の賃金要件", s["wage"])
    r = put_h3(ws, r, "賃金引上げ計画(数値は編集可能)")
    r = put_data_table(ws, r, ["区分", "現行", "計画"], [
        ["事業場内最低賃金(円)", p["wage_before"], p["wage_after1"]],
        ["引上げ額(円)", "—", f"+{p['wage_up_applied']}"],
        ["給与総額", "基準", "+3.5%/年"],
        ["対象者", "全従業員(雇用保険被保険者)", ""],
    ])
    r = put_h3(ws, r, "事業場内最低賃金の推移(円/時)")
    wage_header = r
    r = put_data_table(ws, r, ["時点", "最低賃金"], [
        ["現行", p["wage_before"]], ["計画1年", p["wage_after1"]], ["計画3年", p["wage_after3"]],
    ])
    r = add_bar_chart(ws, r, "事業場内最低賃金の推移(円/時)", wage_header, wage_header + 3, data_col_end=2)
    r = put_h3(ws, r, "社会保険労務士の視点(処遇改善の実効性)")
    r = put_data_table(ws, r, ["観点", "内容"], [
        ["原資の確保", "本事業の省力化で生じた利益を賃上げ原資に充当し、無理のない持続的な賃金引上げを実現する。"],
        ["就業規則の整備", "賃金規程を改定し、事業場内最低賃金の引上げを規程上も明確化。労働条件通知書へ反映する。"],
        ["労働時間の適正化", "省力化で残業を圧縮し、勤務間インターバルと有給取得を促進。働きやすさと処遇の両面で職場を改善する。"],
    ])
    apply_print_setup(ws, r)

    # --- 6. 資金計画・投資回収 ---
    ws = sheet(wb, "6_資金計画")
    r = put_title(ws, 1, "6. 資金計画・投資回収(税理士の視点)", 6)
    r = put_h3(ws, r, "投資内訳(万円・編集可能)")
    inv_header = r
    item_rows = [[k, v] for k, v in p["invItems"]]
    r = put_data_table(ws, r, ["項目", "金額(万円)"], item_rows)
    r = add_pie_chart(ws, r, "投資内訳", inv_header, inv_header + len(item_rows))
    r = put_kv_table(ws, r, [
        ["投資総額(万円)", p["inv"]],
        [f"補助金({s['rate'].split('(')[0]})(万円)", -s["subAmt"]],
        ["自己負担額(万円)", s["self"]],
    ])
    r = put_h3(ws, r, "資金調達")
    r = put_data_table(ws, r, ["区分", "金額(万円)"], [
        ["補助金", s["subAmt"]], ["自己資金", round(s["self"] * 0.5)], ["金融機関借入", s["self"] - round(s["self"] * 0.5)],
    ])
    r = put_h3(ws, r, "累積キャッシュフロー(万円)")
    cf_header = r
    r = put_data_table(ws, r, ["時点", "累積CF"], [[c["y"], c["cum"]] for c in p["cf"]])
    r = add_bar_chart(ws, r, "累積キャッシュフロー(万円)", cf_header, cf_header + len(p["cf"]), data_col_end=2)
    r = put_h3(ws, r, "年間収支")
    r = put_data_table(ws, r, ["区分", "導入時", "1年目", "2年目", "3年目"], [
        ["収益改善効果", "—", p["saveY"], p["saveY"], p["saveY"]],
        ["保守・運用費", "—", -p["opex"], -p["opex"], -p["opex"]],
        ["減価償却(定額・5年)", "—", round(p["inv"] / 5), round(p["inv"] / 5), round(p["inv"] / 5)],
        ["年間純効果", -s["self"], p["netY"], p["netY"], p["netY"]],
    ])
    r = put_note(ws, r, f"投資回収期間 約{p['roi']}年(自己負担ベース)。税務上はソフトウェア等を無形固定資産として計上し定額法で償却。中小企業向け税制の活用余地も税理士と精査する。", span=5)
    apply_print_setup(ws, r)

    # --- 7. 実施体制・スケジュール ---
    ws = sheet(wb, "7_実施体制")
    r = put_title(ws, 1, "7. 実施体制・スケジュール", 7)
    r = put_h3(ws, r, "スケジュール")
    r = put_data_table(ws, r, ["時期", "内容"], [
        ["1か月目", "交付決定後 要件定義"], ["2-3か月", "システム構築・既存連携"],
        ["4か月目", "試験運用・データ移行"], ["5か月目", "スタッフ研修・本稼働"], ["6か月〜", "効果測定・改善運用"],
    ])
    r = put_h3(ws, r, "実施体制")
    r = put_data_table(ws, r, ["役割", "担当"], [
        ["統括責任者", "代表者"], ["現場推進", "店長"], ["システム導入", "ベンダー+IT担当"],
        ["効果測定", "店長+顧問税理士"], ["労務・賃上げ", "顧問社会保険労務士"],
    ])
    r = put_h3(ws, r, "専門家連携")
    r = put_data_table(ws, r, ["専門家", "役割"], [
        ["中小企業診断士", "事業計画の実現性・生産性向上効果を監修"],
        ["税理士", "資金計画・投資回収・減価償却・補助金経理を支援"],
        ["社会保険労務士", "賃上げ計画・就業規則・労務要件を監修"],
        ["ITベンダー", "システム構築・保守・運用定着を担当"],
    ])
    r = put_callout(ws, r, "交付決定前の発注厳禁", "本事業の設備投資・契約は必ず交付決定後に着手する(事前着手は補助対象外)。スケジュールは交付決定を起点として設計している。")
    apply_print_setup(ws, r)

    # --- 8. リスクと対応策 ---
    ws = sheet(wb, "8_リスク")
    r = put_title(ws, 1, "8. リスクと対応策", 8)
    risk_rows = list(p["risks"]) + [
        ["導入が定着しない", "スタッフ研修と段階導入で現場負担を抑え、KPIを可視化して効果を実感させる。ベンダーの伴走支援を契約に含める。"],
        ["効果が計画に届かない", "月次でKPIをモニタリングし、AIの学習・設定を継続改善。四半期ごとにPDCAを回して軌道修正する。"],
    ]
    r = put_data_table(ws, r, ["想定リスク", "対応策"], risk_rows)
    r = put_note(ws, r, "主要リスクはいずれも発生確率・影響度ともに管理可能な範囲にあり、対応策を講じることで事業計画の達成可能性は高いと評価する。")
    apply_print_setup(ws, r)

    # --- 9. 地域・業界への波及効果 ---
    ws = sheet(wb, "9_波及効果")
    r = put_title(ws, 1, "9. 地域・業界への波及効果", 9)
    r = put_lead(ws, r, p["ripple"])
    r = put_h3(ws, r, "3つの波及効果")
    r = put_data_table(ws, r, ["観点", "内容"], [
        ["雇用の質の向上", "生産性向上を原資とした賃上げと働き方改善で、地域の雇用の魅力を高める。"],
        ["モデルの横展開", "本事業のノウハウは同業・近隣他業種へ展開可能で、地域全体のDXを牽引する。"],
        ["顧客利便性の向上", "サービス品質と利便性の向上が、地域住民・来訪者の満足度を高める。"],
    ])
    r = put_callout(ws, r, "中小企業診断士の視点",
                     "本事業は単なる業務効率化にとどまらず、省力化で創出した経営資源を高付加価値活動へ再配分する経営変革である。"
                     "人手不足という構造的制約を乗り越え、小規模事業者が持続的に成長するための投資であり、地域経済の担い手としての基盤強化に資する。")
    apply_print_setup(ws, r)

    # --- 10. 要件チェック・他制度転用 ---
    ws = sheet(wb, "10_要件確認")
    r = put_title(ws, 1, "10. 要件チェック・他制度転用", 10)
    wage_label = "事業場内最低賃金の引上げ" if p["scheme"] == "業務改善" else "賃上げ計画(特例活用時)"
    r = put_h3(ws, r, "補助要件チェックリスト")
    r = put_data_table(ws, r, ["要件", "本計画の対応", "判定"], [
        [f"{s['badge']}の対象事業者", "中小企業・小規模事業者(美容業)に該当", "✔"],
        ["省力化・生産性向上効果", f"月{p['saveH']}時間削減・KPI改善を定量提示", "✔"],
        [wage_label, f"+{p['wage_up_applied']}円の引上げを計画", "✔"],
        ["交付決定後の発注", "スケジュールを交付決定起点で設計", "✔"],
        ["事業計画の実現可能性", "実機プロトタイプ・専門家連携で裏付け", "✔"],
        ["投資回収の妥当性", f"約{p['roi']}年で回収、CF計画を提示", "✔"],
    ])
    if p["scheme"] == "省力化":
        transfer = ("本事業は省力化(生産性向上)効果が明確なため省力化投資補助金(一般型)を主軸とした。"
                    "一方、賃上げを主目的に据え直せば業務改善助成金への転用も可能。その場合は事業場内最低賃金の引上げ額"
                    "(50円以上)を軸に、対象経費を生産性向上に資する設備投資へ組み替える。")
    else:
        transfer = ("本事業は賃上げと業務改善の親和性が高いため業務改善助成金を主軸とした。"
                    "一方、投資規模を拡大し省力化効果を前面に出せば省力化投資補助金(一般型)への転用も可能。"
                    "その場合は補助上限が大きく、大幅賃上げ特例で補助率2/3を狙える。")
    r = put_callout(ws, r, f"本計画のメイン: {s['badge']}", transfer)
    r = put_callout(ws, r, "重要",
                     "本計画書はデモ用のサンプルです。数値(賃金・売上・投資額等)は仮置きであり、実際の申請にあたっては"
                     "①自社の実績データへの差し替え ②最新の公募要領との突合 ③認定支援機関・専門家"
                     "(中小企業診断士・税理士・社会保険労務士)による確認が必須です。補助金の採択を保証するものではありません。")
    r = put_kv_table(ws, r, [["参照した最新公募情報(2026年時点)", "省力化投資補助金(一般型)第6・7回公募 / 業務改善助成金 令和8年度"]])
    apply_print_setup(ws, r)

    return wb


def main():
    with open(os.path.join(HERE, "_plan_data.json"), encoding="utf-8") as f:
        data = json.load(f)
    biz = data["biz"]
    for p in data["plans"]:
        wb = build_plan_workbook(p, biz)
        out = os.path.join(HERE, f"plan-{p['no']:02d}.xlsx")
        wb.save(out)
        print(f"  OK plan-{p['no']:02d}.xlsx")
    print(f"\nDone: {len(data['plans'])} plans exported")


if __name__ == "__main__":
    main()
