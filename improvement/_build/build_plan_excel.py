# -*- coding: utf-8 -*-
"""data/plans/{industry}.json から事業計画書のExcel版 plan-NN.xlsx を生成する。

美容業の plan-NN.xlsx (beauty/_build_plan_excel.py が _plan_data.json から生成)
に対応する、Python生成系業種向けのExcel版。数値はHTML版と同じ計算
(補助見込 = min(上限額, 投資総額×補助率)) をここでも行い、手入力させない。

実行: python3 build_plan_excel.py [industry ...]   省略時は plans/ にある全業種
（beauty は外部プロジェクト連携のためスキップ。HTML側と同じガード）
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
IMPROVEMENT = os.path.abspath(os.path.join(HERE, ".."))

INDUSTRY_ACCENT = {
    "beauty": "8A5A2B", "food": "A4453C", "lodging": "3E6B7A",
    "manufacturing": "4A5A7A", "realestate": "3E6B4F", "education": "7A5A8A",
}
INDUSTRY_LABEL = {
    "beauty": "美容業", "food": "飲食業", "lodging": "宿泊業",
    "manufacturing": "製造業", "realestate": "不動産業", "education": "教育・学習支援業",
}
EXTERNAL_INDUSTRIES = {"beauty"}

THIN = Side(style="thin", color="C9C2B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _header(ws, row, texts, accent):
    for col, t in enumerate(texts, start=1):
        c = ws.cell(row=row, column=col, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
    return row + 1


def _rows(ws, row, rows, bold_first_col=False):
    for r in rows:
        for col, v in enumerate(r, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.alignment = WRAP
            if bold_first_col and col == 1:
                c.font = Font(bold=True)
        row += 1
    return row


def _title(ws, text, accent, span=3):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=accent)
    return 3


def build_xlsx(pl, accent, industry_label, out_path):
    inv = pl["investment"]
    sub = pl["subsidy"]
    total = inv["total"]
    applied = round(min(sub["cap"], total * sub["rate"]), 1)
    if applied == int(applied):
        applied = int(applied)
    self_pay = round(total - applied, 1)
    if self_pay == int(self_pay):
        self_pay = int(self_pay)
    eff = pl["effect"]
    monthly = eff.get("monthlySaving", {}).get("amount")
    payback_m = round(self_pay / monthly) if monthly else None

    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1. 計画概要
    ws = _sheet(wb, "計画概要", [22, 60])
    r = _title(ws, f"事業計画書 PLAN {pl['no']:02d}（{industry_label}・モデルケース）", accent)
    rows = [("計画名", pl["title"]), ("サブタイトル", pl["subtitle"]),
            ("分類", pl["category"]["name"])]
    rows += [tuple(l) for l in pl["model"]["lines"]]
    rows += [("投資総額", f"{total}万円（{inv['unit']}）"),
             ("活用制度", sub["label"]), ("補助率", sub["rateText"]), ("上限額", sub["capText"]),
             ("補助見込", f"{applied}万円 = min(上限{sub['cap']}万円, {total}万円×{sub['rate']:g})"),
             ("自己負担", f"{self_pay}万円")]
    if payback_m:
        rows.append(("投資回収（試算）", f"約{payback_m}ヶ月"))
    rows.append(("本計画の狙い", pl["solution"]["aim"]))
    r = _rows(ws, r, rows, bold_first_col=True)
    ws.cell(row=r + 1, column=1, value="※ 数値は公開情報・業界目安に基づくモデルケースの仮置き。実申請時は自社実績に差し替えること。").font = Font(size=9, color="8A8178")

    # ---- 2. 費用・補助金
    ws = _sheet(wb, "費用・補助金", [40, 16, 40])
    r = _title(ws, "費用内訳と補助金試算", accent)
    r = _header(ws, r, ["費目", "金額(万円)", "備考"], accent)
    r = _rows(ws, r, [(l, v, "") for l, v in inv["lines"]])
    r = _rows(ws, r, [("合計", total, inv["unit"])], bold_first_col=True)
    r += 1
    r = _header(ws, r, ["補助金試算", "金額(万円)", "根拠"], accent)
    r = _rows(ws, r, [
        ("補助見込", applied, f"min(上限{sub['cap']}万円, 総額{total}万円×補助率{sub['rate']:g})"),
        ("自己負担", self_pay, "投資総額 − 補助見込"),
    ], bold_first_col=True)
    r += 1
    r = _rows(ws, r, [("費用の根拠", inv["basisNote"], "")], bold_first_col=True)
    ws.cell(row=r + 1, column=1,
            value="※ 交付決定前に発注・契約・支払いを行った経費は原則補助対象外。公募要領の最新版を必ず確認。").font = Font(size=9, color="B3372C")

    # ---- 3. 効果・KPI
    ws = _sheet(wb, "効果・KPI", [34, 14, 14, 14])
    r = _title(ws, "効果試算とKPI", accent)
    r = _header(ws, r, ["指標（モデルケース試算）", "導入前", "導入後", ""], accent)
    m = eff["model"]
    ba_rows = [(m["label"], m["before"], m["after"], m.get("basis", ""))]
    ba_rows += [(e["label"], e["before"], e["after"], "") for e in eff.get("modelExtras", [])]
    r = _rows(ws, r, ba_rows)
    if monthly:
        r = _rows(ws, r, [("削減効果額（試算）", "", f"{monthly}万円/月", eff["monthlySaving"]["assumption"])])
    r += 1
    r = _header(ws, r, ["KPI", "現状", "目標", "達成時期"], accent)
    r = _rows(ws, r, [(k["name"], k["base"], k["target"], k["when"]) for k in pl["kpi"]])
    r += 1
    r = _header(ws, r, ["効果データの出典（参考値）", "", "", ""], accent)
    r = _rows(ws, r, [(b, "", "", "") for b in eff["sourceBullets"]])

    # ---- 4. スケジュール
    ws = _sheet(wb, "スケジュール", [40, 12, 12])
    r = _title(ws, "導入スケジュール（想定）", accent)
    r = _header(ws, r, ["実施内容", "開始月", "終了月"], accent)
    r = _rows(ws, r, [(s["label"], s["m"][0], s["m"][1]) for s in pl["schedule"]])
    r += 1
    r = _rows(ws, r, [("補足", pl["scheduleNote"], "")], bold_first_col=True)

    # ---- 5. リスク・アクション
    ws = _sheet(wb, "リスク・アクション", [45, 55])
    r = _title(ws, "想定リスクと次のアクション", accent)
    r = _header(ws, r, ["想定リスク", "対応策"], accent)
    r = _rows(ws, r, [(x["risk"], x["counter"]) for x in pl["risks"]])
    r += 1
    r = _header(ws, r, ["次のアクション", ""], accent)
    r = _rows(ws, r, [(a, "") for a in pl["actions"]])
    ws.cell(row=r + 1, column=1,
            value="※ 本計画書はモデルケースの試算・提案であり、採択・効果を保証するものではない。").font = Font(size=9, color="8A8178")

    wb.save(out_path)


def main():
    industries = sys.argv[1:]
    if not industries:
        pdir = os.path.join(DATA, "plans")
        industries = [f[:-5] for f in os.listdir(pdir) if f.endswith(".json")] if os.path.isdir(pdir) else []
    for ik in industries:
        if ik in EXTERNAL_INDUSTRIES:
            print(f"SKIP: {ik} は外部プロジェクト連携のため build_plan_excel.py では生成しません")
            continue
        with open(os.path.join(DATA, "plans", ik + ".json"), encoding="utf-8") as f:
            plans = json.load(f)
        outdir = os.path.join(IMPROVEMENT, ik)
        for pl in plans:
            build_xlsx(pl, INDUSTRY_ACCENT[ik], INDUSTRY_LABEL[ik],
                       os.path.join(outdir, f"plan-{pl['no']:02d}.xlsx"))
        print(f"OK: {ik} -> {len(plans)} xlsx")


if __name__ == "__main__":
    main()
